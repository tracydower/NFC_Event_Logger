"""
NFC → Dropbox spreadsheet logger
=================================

A tiny web service that receives an event from the iPhone Shortcuts
("Automations") app when you tap an NFC tag, and appends a row to a
spreadsheet (.xlsx) stored in your Dropbox.

Flow:
    Tap NFC tag  ->  Shortcuts automation  ->  HTTP POST to /log  ->  this app
                                                                        |
                                                        appends a row to Dropbox xlsx

Each row that gets logged contains:
    - timestamp_utc    (when the tap happened, in UTC)
    - timestamp_local  (same moment, in your local timezone)
    - tag              (the tag name you send from Shortcuts, e.g. "front door")

Environment variables (see .env.example):
    API_TOKEN               A secret you make up. Shortcuts must send it so
                            strangers can't write to your sheet.
    DROPBOX_APP_KEY         From the Dropbox App Console.
    DROPBOX_APP_SECRET      From the Dropbox App Console.
    DROPBOX_REFRESH_TOKEN   Produced once by running get_refresh_token.py.
    DROPBOX_FILE_PATH       Where to store the sheet in Dropbox, e.g. /nfc_log.xlsx
    TIMEZONE                IANA tz name for the local column, e.g. America/Chicago
"""

import io
import os
from datetime import datetime, timezone

try:
    # Python 3.9+ standard library
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

import dropbox
from dropbox.files import WriteMode
from flask import Flask, jsonify, request
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Configuration (read once at startup)
# ---------------------------------------------------------------------------
API_TOKEN = os.environ.get("API_TOKEN", "")
DROPBOX_APP_KEY = os.environ.get("DROPBOX_APP_KEY", "")
DROPBOX_APP_SECRET = os.environ.get("DROPBOX_APP_SECRET", "")
DROPBOX_REFRESH_TOKEN = os.environ.get("DROPBOX_REFRESH_TOKEN", "")
DROPBOX_FILE_PATH = os.environ.get("DROPBOX_FILE_PATH", "/nfc_log.xlsx")
TIMEZONE = os.environ.get("TIMEZONE", "America/Chicago")

HEADERS = ["timestamp_utc", "timestamp_local", "tag"]

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Dropbox helpers
# ---------------------------------------------------------------------------
def get_dropbox_client():
    """Return an authenticated Dropbox client.

    We use a refresh token so the app can keep working forever without you
    having to log in again. The SDK automatically trades the refresh token
    for a fresh short-lived access token as needed.
    """
    missing = [
        name
        for name, value in {
            "DROPBOX_APP_KEY": DROPBOX_APP_KEY,
            "DROPBOX_APP_SECRET": DROPBOX_APP_SECRET,
            "DROPBOX_REFRESH_TOKEN": DROPBOX_REFRESH_TOKEN,
        }.items()
        if not value
    ]
    if missing:
        raise RuntimeError(
            "Missing Dropbox config: " + ", ".join(missing) +
            ". Set these environment variables (see README)."
        )

    return dropbox.Dropbox(
        oauth2_refresh_token=DROPBOX_REFRESH_TOKEN,
        app_key=DROPBOX_APP_KEY,
        app_secret=DROPBOX_APP_SECRET,
    )


def download_workbook(dbx):
    """Download the existing spreadsheet, or create a fresh one with headers."""
    try:
        _metadata, response = dbx.files_download(DROPBOX_FILE_PATH)
        return load_workbook(io.BytesIO(response.content))
    except dropbox.exceptions.ApiError as err:
        # File doesn't exist yet -> start a new workbook with a header row.
        is_not_found = (
            err.error.is_path()
            and err.error.get_path().is_not_found()
        )
        if is_not_found:
            wb = Workbook()
            ws = wb.active
            ws.title = "log"
            ws.append(HEADERS)
            return wb
        raise


def upload_workbook(dbx, wb):
    """Upload the workbook back to Dropbox, overwriting the old copy."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    dbx.files_upload(
        buffer.read(),
        DROPBOX_FILE_PATH,
        mode=WriteMode.overwrite,
    )


def append_row(tag, when_utc):
    """Append a single (timestamp, tag) row to the Dropbox spreadsheet."""
    dbx = get_dropbox_client()
    wb = download_workbook(dbx)
    ws = wb.active

    # Build the local-time string.
    if ZoneInfo is not None:
        try:
            local_dt = when_utc.astimezone(ZoneInfo(TIMEZONE))
        except Exception:
            local_dt = when_utc
    else:
        local_dt = when_utc

    ws.append([
        when_utc.strftime("%Y-%m-%d %H:%M:%S"),
        local_dt.strftime("%Y-%m-%d %H:%M:%S"),
        tag,
    ])
    upload_workbook(dbx, wb)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.get("/")
def health():
    """Simple health check so you can confirm the service is running."""
    return jsonify({"status": "ok", "service": "nfc-dropbox-logger"})


@app.post("/log")
def log_event():
    """Receive an NFC event from Shortcuts and append a row to Dropbox.

    Accepts either JSON body or form fields:
        tag    (required)  the tag name, e.g. "front door"
        token  (required)  must equal API_TOKEN
    The token may also be sent as a query string (?token=...) or an
    "Authorization: Bearer <token>" header, whichever is easiest in Shortcuts.
    """
    # Pull values from JSON, form, or query string — whichever is present.
    data = request.get_json(silent=True) or {}

    def field(name):
        return (
            data.get(name)
            or request.form.get(name)
            or request.args.get(name)
        )

    # Token can arrive several ways; accept the most convenient.
    token = field("token")
    auth_header = request.headers.get("Authorization", "")
    if not token and auth_header.lower().startswith("bearer "):
        token = auth_header[7:].strip()

    if not API_TOKEN:
        return jsonify({"error": "server missing API_TOKEN config"}), 500
    if token != API_TOKEN:
        return jsonify({"error": "unauthorized"}), 401

    tag = field("tag")
    if not tag:
        return jsonify({"error": "missing 'tag'"}), 400

    when_utc = datetime.now(timezone.utc)

    try:
        append_row(str(tag), when_utc)
    except Exception as exc:  # noqa: BLE001 - report any failure to the caller
        return jsonify({"error": str(exc)}), 500

    return jsonify({
        "status": "logged",
        "tag": tag,
        "timestamp_utc": when_utc.strftime("%Y-%m-%d %H:%M:%S"),
    })


if __name__ == "__main__":
    # Local development server. In production, gunicorn runs the app (see Procfile).
    port = int(os.environ.get("PORT", "8000"))
    app.run(host="0.0.0.0", port=port)
