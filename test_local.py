"""
Local smoke test — no Dropbox account needed.
=============================================

This fakes the Dropbox part so you can confirm the web service and the
row-appending logic work on your own machine before deploying.

Usage:
    pip install -r requirements.txt
    python test_local.py

It spins up the app in "fake Dropbox" mode, sends a couple of test taps,
and writes the resulting spreadsheet to ./local_test_output.xlsx so you can
open it and see the rows.
"""

import io
import os

os.environ.setdefault("API_TOKEN", "test-token")
# Dummy Dropbox creds so the app imports cleanly; we monkeypatch the client.
os.environ.setdefault("DROPBOX_APP_KEY", "x")
os.environ.setdefault("DROPBOX_APP_SECRET", "x")
os.environ.setdefault("DROPBOX_REFRESH_TOKEN", "x")

import app as service  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

# In-memory stand-in for the file that would live on Dropbox.
_fake_store = {"bytes": None}


class FakeDbx:
    def files_download(self, path):
        if _fake_store["bytes"] is None:
            raise service.dropbox.exceptions.ApiError(
                request_id="1",
                error=None,
                user_message_text=None,
                user_message_locale=None,
            )
        resp = type("R", (), {"content": _fake_store["bytes"]})()
        return (None, resp)

    def files_upload(self, data, path, mode=None):
        _fake_store["bytes"] = data


# Patch: use our fake client, and treat "no file yet" as not-found.
service.get_dropbox_client = lambda: FakeDbx()


def _fake_download_workbook(dbx):
    if _fake_store["bytes"] is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "log"
        ws.append(service.HEADERS)
        return wb
    return load_workbook(io.BytesIO(_fake_store["bytes"]))


service.download_workbook = _fake_download_workbook


def main():
    client = service.app.test_client()

    # Health check
    r = client.get("/")
    assert r.status_code == 200, r.data
    print("health:", r.get_json())

    # Wrong token should be rejected
    r = client.post("/log", json={"tag": "front door", "token": "WRONG"})
    assert r.status_code == 401, r.data
    print("bad token correctly rejected:", r.get_json())

    # Missing tag should be rejected
    r = client.post("/log", json={"token": "test-token"})
    assert r.status_code == 400, r.data
    print("missing tag correctly rejected:", r.get_json())

    # Two good taps
    for tag in ["front door", "gym"]:
        r = client.post("/log", json={"tag": tag, "token": "test-token"})
        assert r.status_code == 200, r.data
        print("logged:", r.get_json())

    # Save the resulting sheet so you can open it
    with open("local_test_output.xlsx", "wb") as f:
        f.write(_fake_store["bytes"])

    wb = load_workbook("local_test_output.xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print("\nSpreadsheet contents:")
    for row in rows:
        print("  ", row)

    assert rows[0] == tuple(service.HEADERS)
    assert rows[1][2] == "front door"
    assert rows[2][2] == "gym"
    print("\nAll checks passed. Open local_test_output.xlsx to see the sheet.")


if __name__ == "__main__":
    main()
